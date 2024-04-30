from pathlib import Path
from datetime import date
from datetime import datetime
import click
import sys
import yaml
from typing import Any, IO
from rich import print
import os
import json
# from .myloadenv import load_env
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class FragenConv:

    def __init__(self, main_input, template, outfile):
        self._main_input = Path(main_input)
        self._today = date.today().strftime("%d.%m.%Y")
        with open(self._main_input, 'r', encoding="utf8") as f:
            self._data = yaml.load(f, Loader=yaml.CLoader)
        self._path = self._main_input.parent
        self._cwd = Path.cwd()

        self._outfile = Path(outfile)
        self._template = Path(template)
        self._load_fragen()

    @property
    def data(self):
        return self._data

    def _load_fragen(self):
        for le in self._data['lerneinheiten']:
            fragen_datei = self._path / le['datei']
            with open(fragen_datei, 'r', encoding="utf8") as f:
                temp = yaml.load(f, Loader=yaml.CLoader)
                le['fragenpool'] = temp

    def _write_dekra(self, wb):
        bp = wb["Fragebogen Vorlage"]
        r = start_row = 3
        c = start_col = 2
        for le in self._data["lerneinheiten"]:
            fragenpool = le['fragenpool']
            if fragenpool['fragen'] is None:
                continue
            for frage in fragenpool['fragen']:
                if not frage['freigegeben'] == 'y':
                    continue
                # Ordnerhierarchie in EvaExam
                bp.cell(row=r, column=c).value = "TPSSE"
                bp.cell(row=r, column=c + 1).value = "T.P.S.S.E."
                bp.cell(row=r, column=c + 2).value = fragenpool['thema']
                # Fragenparameter
                bp.cell(row=r, column=c + 3).value = frage['id']
                bp.cell(row=r, column=c + 4).value = "Multiple-Choice (Korrekte Kombination)"  # F-Form
                #bp.cell(row=r, column=c + 5).value =  # Schwierigkeitsgrad
                bp.cell(row=r, column=c + 6).value = frage['text']  # Frage
                # Antwortmöglichkeiten
                for o_index, optionen in enumerate(frage['optionen']):
                    bp.cell(row=r, column=c + 7 + o_index).value = optionen['text']
                    bp.cell(row=r, column=c + 13 + o_index).value = int(optionen['richtig'] == 'y')
                r = r + 1

    def write_xls(self):
        wb = load_workbook(self._template)
        self._write_dekra(wb)
        wb.save(self._outfile)


@click.command()
@click.option('--fmain', envvar='FRAGENCONV_MAIN', help="Main-Datei für Prüfungsfragen (FRAGENCONV_MAIN)")
@click.option('--template', envvar='FRAGENCONV_TEMPLATE', help="Main-Datei für Prüfungsfragen (FRAGENCONV_TEMPLATE)")
@click.option('--out', envvar='FRAGENCONV_OUT', help="Main-Datei für Prüfungsfragen (FRAGENCONV_OUT)")
def run(fmain, template, out):
    converter = FragenConv(fmain, template, out)
    converter.write_xls()


def main():
    # load_env()
    run()


if __name__ == '__main__':
    main()
